"""
Enhanced Location Pins Template Generator

Creates Excel templates with latitude/longitude columns for KML import compatibility.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


def create_enhanced_template(output_path: str = "location_pins_enhanced_template.xlsx"):
    """
    Create an enhanced location pins template with latitude/longitude columns.
    
    Args:
        output_path: Path where the template will be saved
    """
    
    # Define the enhanced column structure
    columns = [
        "Location Name",
        "Street Address", 
        "City",
        "State",
        "ZIP/Postal Code",
        "Latitude",
        "Longitude", 
        "Electrification Candidates",
        "Category Name"
    ]
    
    # Create sample data with explanations
    sample_data = [
        {
            "Location Name": "Example Distribution Center",
            "Street Address": "123 Industrial Blvd",
            "City": "Commerce City",
            "State": "CO",
            "ZIP/Postal Code": "80022",
            "Latitude": "39.8083",
            "Longitude": "-104.9342",
            "Electrification Candidates": "5",
            "Category Name": "Distribution Centers"
        },
        {
            "Location Name": "Sample Logistics Hub",
            "Street Address": "",
            "City": "Denver",
            "State": "CO", 
            "ZIP/Postal Code": "80202",
            "Latitude": "",
            "Longitude": "",
            "Electrification Candidates": "3",
            "Category Name": "Logistics Hubs"
        },
        {
            "Location Name": "Test Location - Coordinates Only",
            "Street Address": "",
            "City": "",
            "State": "",
            "ZIP/Postal Code": "",
            "Latitude": "39.7392",
            "Longitude": "-104.9903",
            "Electrification Candidates": "1",
            "Category Name": "Test Locations"
        }
    ]
    
    # Create DataFrame
    df = pd.DataFrame(sample_data, columns=columns)
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Location Pins"
    
    # Add instructions worksheet
    instructions_ws = wb.create_sheet("Instructions", 0)
    
    # Write instructions
    instructions = [
        ["Enhanced Location Pins Template", ""],
        ["", ""],
        ["This template supports both address-based and coordinate-based locations:", ""],
        ["", ""],
        ["REQUIRED COLUMNS:", ""],
        ["• Location Name", "Descriptive name for the location"],
        ["• Electrification Candidates", "Number of vehicles/candidates (integer)"],
        ["• Category Name", "Category for grouping locations"],
        ["", ""],
        ["LOCATION DATA (use one or both):", ""],
        ["Option 1 - Address-based:", ""],
        ["• Street Address", "Physical street address (optional)"],
        ["• City", "City name (optional)"],
        ["• State", "State abbreviation (optional)"],
        ["• ZIP/Postal Code", "ZIP or postal code (helpful for geocoding)"],
        ["", ""],
        ["Option 2 - Coordinate-based:", ""],
        ["• Latitude", "Decimal degrees (e.g., 39.7392)"],
        ["• Longitude", "Decimal degrees (e.g., -104.9903)"],
        ["", ""],
        ["NOTES:", ""],
        ["• If coordinates are provided, they take priority over addresses"],
        ["• If only partial address info, system will try to geocode"],
        ["• If geocoding fails, system falls back to state centroid"],
        ["• Delete the sample rows and add your own data"],
        ["• Save as Excel (.xlsx) or CSV (.csv) format"],
        ["", ""],
        ["COORDINATE FORMATS SUPPORTED:", ""],
        ["• Decimal degrees (preferred): 39.7392, -104.9903", ""],
        ["• Must be in WGS84 coordinate system", ""],
        ["• Latitude range: -90 to 90", ""],
        ["• Longitude range: -180 to 180", ""],
    ]
    
    for row_idx, instruction_row in enumerate(instructions, 1):
        if len(instruction_row) == 2:
            instruction, description = instruction_row
        else:
            instruction = instruction_row[0] if instruction_row else ""
            description = ""
            
        ws_instructions = instructions_ws.cell(row=row_idx, column=1, value=instruction)
        if description:
            instructions_ws.cell(row=row_idx, column=2, value=description)
        
        # Style the headers
        if instruction and ":" in instruction and not description:
            ws_instructions.font = Font(bold=True, color="0066CC")
        elif instruction and not description and not instruction.startswith("•"):
            ws_instructions.font = Font(bold=True, size=14, color="FF0000")
    
    # Adjust column widths for instructions
    instructions_ws.column_dimensions['A'].width = 25
    instructions_ws.column_dimensions['B'].width = 50
    
    # Write data to main worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Style the header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Adjust column widths
    column_widths = [25, 30, 15, 8, 15, 12, 12, 22, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # Add data validation for coordinates (basic)
    from openpyxl.worksheet.datavalidation import DataValidation
    
    # Latitude validation (-90 to 90)
    lat_validation = DataValidation(
        type="decimal", 
        operator="between",
        formula1=-90,
        formula2=90,
        allow_blank=True,
        errorTitle="Invalid Latitude",
        error="Latitude must be between -90 and 90 degrees"
    )
    ws.add_data_validation(lat_validation)
    lat_validation.add(f'F2:F1000')  # Latitude column
    
    # Longitude validation (-180 to 180)
    lon_validation = DataValidation(
        type="decimal",
        operator="between", 
        formula1=-180,
        formula2=180,
        allow_blank=True,
        errorTitle="Invalid Longitude",
        error="Longitude must be between -180 and 180 degrees"
    )
    ws.add_data_validation(lon_validation)
    lon_validation.add(f'G2:G1000')  # Longitude column
    
    # Electrification Candidates validation (positive integers)
    candidates_validation = DataValidation(
        type="whole",
        operator="greaterThan",
        formula1=0,
        allow_blank=False,
        errorTitle="Invalid Candidates Count",
        error="Electrification Candidates must be a positive integer"
    )
    ws.add_data_validation(candidates_validation)
    candidates_validation.add(f'H2:H1000')  # Electrification Candidates column
    
    # Save the workbook
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    template_path = create_enhanced_template()
    print(f"Enhanced location pins template created: {template_path}")
